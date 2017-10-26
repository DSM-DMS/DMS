import openpyxl
from flask import send_from_directory
from flask_jwt import current_identity, jwt_required
from flask_restful_swagger_2 import Resource, swagger

from db.models.account import AdminModel, StudentModel

from . import goingout_doc


class Goingout(Resource):
    @swagger.doc(goingout_doc.GOINGOUT_GET)
    @jwt_required()
    def get(self):
        """
        외출신청 엑셀 다운로드
        """
        admin = AdminModel.objects(id=current_identity)

        if not admin:
            return '', 403

        wb = openpyxl.load_workbook('명렬표.xlsx')
        ws = wb.active

        for row in map(str, range(3, 68)):
            for column1, column2, column3 in zip(['B', 'F', 'J', 'N'], ['D', 'H', 'L', 'P'], ['E', 'I', 'M', 'Q']):
                student = StudentModel.objects(number=ws[column1+row]).first()

                if not student:
                    continue

                sat = '토요 외출' if student.goingout_applies.on_saturday else ''
                sun = '일요 외출' if student.goingout_applies.on_sunday else ''
                ws[column2+row] = sat
                ws[column3+row] = sun

        wb.save('명렬표.xlsx')
        return send_from_directory('주소오오오', '명렬표.xlsx'), 200
