from flask import send_from_directory
from flask_restful_swagger_2 import Resource, swagger
from flask_jwt import current_identity, jwt_required
import openpyxl

from db.models.account import AdminModel, StudentModel
from . import stay_doc


class Stay(Resource):
    @swagger.doc(stay_doc.STAY_GET)
    @jwt_required()
    def get(self):
        """
        잔류신청 엑셀 다운로드
        """
        admin = AdminModel.objects(id=current_identity)

        if not admin:
            return '', 403

        wb = openpyxl.load_workbook('명렬표.xlsx')
        ws = wb.active

        for row in map(str, range(3, 68)):
            for column1, column2 in zip(['B', 'F', 'J', 'N'], ['D', 'H', 'L', 'P']):
                student = StudentModel.objects(number=ws[column1+row]).first()
                status = ''
                value = student.stay_apply.value

                if not student:
                    continue

                if value == '1':
                    status = '금요 귀가'
                elif value == '2':
                    status = '토요 귀가'
                elif value == '3':
                    status = '토요 귀사'
                elif value == '4':
                    status = '잔류'
                ws[column2+row] = status

        wb.save('명렬표.xlsx')
        return send_from_directory('주소오오오', '명렬표.xlsx'), 200
