import openpyxl

from flask import send_from_directory
from flask_jwt_extended import get_jwt_identity, jwt_required
from flask_restful_swagger_2 import Resource, swagger

from db.models.account import AdminModel, StudentModel
from routes.api.admin.apply import extension_doc


class Extension(Resource):
    uri = '/admin/extension'

    @swagger.doc(extension_doc.EXTENSION_GET)
    @jwt_required
    def get(self):
        """
        연장신청 엑셀 다운로드
        """
        admin = AdminModel.objects(id=get_jwt_identity())

        if not admin:
            return '', 403

        wb = openpyxl.load_workbook('연장 명렬표.xlsx')
        ws = wb.active

        for row in map(str, range(3, 68)):
            for column1, column2 in zip(['B', 'F', 'J', 'N'], ['D', 'H', 'L', 'P']):
                if ws[column1+row].value == '학번':
                    continue

                number = int(ws[column1+row].value or 0)
                student = StudentModel.objects(number=number).first()
                class_ = student.extension_apply.class_ or 0

                if class_ == 0:
                    status = ''
                elif class_ == 1:
                    status = '가온실'
                elif class_ == 2:
                    status = '나온실'
                elif class_ == 3:
                    status = '다온실'
                elif class_ == 4:
                    status = '라온실'
                elif class_ == 5:
                    status = '3층 독서실'
                elif class_ == 6:
                    status = '4층 독서실'
                elif class_ == 7:
                    status = '5층 독서실'

                ws[column2+row] = status

        wb.save('명렬표.xlsx')
        
        return send_from_directory('.', '연장 명렬표.xlsx')
